"""
Builds a single Excel workbook comparing all 16 new_way/ combos (8
architectures x 2 tissue types) -- headline metrics, per-country
stratified metrics, and each model's best Optuna hyperparameters -- read
directly from classification/new_way/Output/logs/*_study_summary.json (the
real, committed 2026-08-11 Kaggle run; see
classification/.project_memory/08_new_way_architecture_roster.md).

Combo discovery is dynamic (globs *_study_summary.json), not a hardcoded
list, matching this project's existing organize_and_compare.py precedent
(classification/v2_clean_scripts/organize_and_compare.py).

Usage: python build_comparison_excel.py
"""
import json
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

THIS_DIR = Path(__file__).resolve().parent
VERSION_DIR = THIS_DIR.parent  # .../new_way/Output/version1
LOGS_DIR = VERSION_DIR / "logs"
OUTPUT_XLSX = THIS_DIR / "new_way_model_comparison.xlsx"

# Architecture metadata (family/tier/params), from classification/
# .project_memory/08_new_way_architecture_roster.md's verified-parameter-
# count table -- not re-derived here, cited from that record.
ARCH_META = {
    "efficientnet_b3": {"family": "CNN", "tier": "Light", "params_m": 12.23},
    "efficientnet_b4": {"family": "CNN", "tier": "Light", "params_m": 19.34},
    "regnet_y_16gf": {"family": "CNN", "tier": "Medium", "params_m": 83.59},
    "convnext_base": {"family": "CNN", "tier": "Medium", "params_m": 88.59},
    "convnext_large": {"family": "CNN", "tier": "Heavy", "params_m": 197.77},
    "maxvit_t": {"family": "Hybrid", "tier": "Light", "params_m": 30.92},
    "maxvit_small": {"family": "Hybrid", "tier": "Medium", "params_m": 68.16},
    "coatnet_3": {"family": "Hybrid", "tier": "Heavy", "params_m": 163.64},
}
TISSUE_TYPES = ["palpebral", "forniceal_palpebral"]


def parse_arch_and_tissue(model_name: str) -> tuple:
    stem = model_name[: -len("_new_way")] if model_name.endswith("_new_way") else model_name
    for tissue in sorted(TISSUE_TYPES, key=len, reverse=True):  # forniceal_palpebral before palpebral
        if stem.endswith(f"_{tissue}"):
            return stem[: -len(f"_{tissue}")], tissue
    raise ValueError(f"Could not parse tissue type from model_name={model_name!r}")


def load_combos() -> list:
    rows = []
    for path in sorted(LOGS_DIR.glob("*_study_summary.json")):
        with open(path, encoding="utf-8") as f:
            summary = json.load(f)

        model_name = summary["model_name"]
        arch_key, tissue = parse_arch_and_tissue(model_name)
        meta = ARCH_META[arch_key]

        by_country = summary["best_val_metrics_by_country"]
        overall = by_country["overall"]
        india = by_country.get("India", {})
        italy = by_country.get("Italy", {})
        params = summary["best_params"]

        def cm_parts(bucket):
            cm = bucket.get("confusion_matrix")
            if cm is None:
                return (None, None, None, None)
            (tn, fp), (fn, tp) = cm
            return (tn, fp, fn, tp)

        tn, fp, fn, tp = cm_parts(overall)
        tn_i, fp_i, fn_i, tp_i = cm_parts(india)
        tn_it, fp_it, fn_it, tp_it = cm_parts(italy)

        rows.append(
            {
                "model_name": model_name,
                "arch_key": arch_key,
                "tissue_type": tissue,
                "family": meta["family"],
                "tier": meta["tier"],
                "params_m": meta["params_m"],
                "n_trials_run": summary["n_trials_run"],
                "best_trial_number": summary["best_trial_number"],
                "learning_rate": params.get("learning_rate"),
                "weight_decay": params.get("weight_decay"),
                "dropout_rate": params.get("dropout_rate"),
                "f1": overall.get("f1"),
                "accuracy": overall.get("accuracy"),
                "precision": overall.get("precision"),
                "recall": overall.get("recall"),
                "specificity": overall.get("specificity"),
                "balanced_accuracy": overall.get("balanced_accuracy"),
                "auc_overall": overall.get("auc"),
                "n_overall": overall.get("n"),
                "tn": tn, "fp": fp, "fn": fn, "tp": tp,
                "n_india": india.get("n"),
                "accuracy_india": india.get("accuracy"),
                "precision_india": india.get("precision"),
                "recall_india": india.get("recall"),
                "specificity_india": india.get("specificity"),
                "balanced_accuracy_india": india.get("balanced_accuracy"),
                "f1_india": india.get("f1"),
                "auc_india": india.get("auc"),
                "tn_india": tn_i, "fp_india": fp_i, "fn_india": fn_i, "tp_india": tp_i,
                "n_italy": italy.get("n"),
                "accuracy_italy": italy.get("accuracy"),
                "precision_italy": italy.get("precision"),
                "recall_italy": italy.get("recall"),
                "specificity_italy": italy.get("specificity"),
                "balanced_accuracy_italy": italy.get("balanced_accuracy"),
                "f1_italy": italy.get("f1"),
                "auc_italy": italy.get("auc"),
                "tn_italy": tn_it, "fp_italy": fp_it, "fn_italy": fn_it, "tp_italy": tp_it,
                "timestamp_utc": summary.get("timestamp_utc"),
            }
        )
    return rows


# --------------------------------------------------------------------------
# Styling
# --------------------------------------------------------------------------
FONT_NAME = "Arial"
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=10)
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
BODY_FONT = Font(name=FONT_NAME, size=10)
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=14)
SUBTITLE_FONT = Font(name=FONT_NAME, italic=True, size=9, color="555555")
FLAG_FILL = PatternFill("solid", fgColor="FFC7CE")  # light red -- flags the known collapsed combo
THIN_BORDER = Border(*(Side(style="thin", color="D9D9D9") for _ in range(4)))
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")

METRIC_FMT = "0.0000"
PCT_FMT = "0.00%"
INT_FMT = "0"
SCI_FMT = "0.000E+00"


def style_header_row(ws, row: int, n_cols: int) -> None:
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def autofit(ws, widths: dict) -> None:
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


def write_table(ws, start_row: int, headers: list, data_rows: list, col_formats: dict) -> int:
    """Writes a header row + data rows starting at start_row, col A.
    col_formats: 1-indexed column number -> number_format string.
    Returns the row number of the last data row."""
    for j, h in enumerate(headers, start=1):
        ws.cell(row=start_row, column=j, value=h)
    style_header_row(ws, start_row, len(headers))

    for i, row in enumerate(data_rows):
        r = start_row + 1 + i
        for j, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=j, value=val)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            if j in col_formats:
                cell.number_format = col_formats[j]
            cell.alignment = LEFT if j <= 3 else CENTER
    return start_row + len(data_rows)


# --------------------------------------------------------------------------
# Build workbook
# --------------------------------------------------------------------------
def build_workbook(rows: list) -> openpyxl.Workbook:
    rows = sorted(rows, key=lambda r: (r["f1"] is None, -(r["f1"] or 0)))

    wb = openpyxl.Workbook()

    # ---- Sheet 1: Summary ----
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "new_way/ -- 16-Combo Model Comparison"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = (
        "8 architectures (5 CNN + 3 Hybrid) x 2 tissue types, 12-trial Optuna search each. "
        "Source: classification/new_way/Output/logs/*_study_summary.json (Kaggle run committed 2026-08-11)."
    )
    ws["A2"].font = SUBTITLE_FONT

    headers = [
        "Rank", "Model", "Tissue Type", "Family", "Tier", "Params (M)",
        "N Trials", "Best Trial #",
        "Learning Rate", "Weight Decay", "Dropout Rate",
        "F1", "Accuracy", "Precision", "Recall", "Specificity", "Balanced Accuracy", "AUC",
        "AUC (India)", "AUC (Italy)", "India/Italy AUC Gap",
    ]
    start_row = 4
    data_rows = []
    for r in rows:
        data_rows.append(
            [
                None,  # Rank filled via formula below
                r["model_name"], r["tissue_type"], r["family"], r["tier"], r["params_m"],
                r["n_trials_run"], r["best_trial_number"],
                r["learning_rate"], r["weight_decay"], r["dropout_rate"],
                r["f1"], r["accuracy"], r["precision"], r["recall"], r["specificity"], r["balanced_accuracy"], r["auc_overall"],
                r["auc_india"], r["auc_italy"], None,  # AUC gap filled via formula below
            ]
        )

    col_formats = {
        6: "0.00", 9: SCI_FMT, 10: SCI_FMT, 11: "0.00",
        12: METRIC_FMT, 13: METRIC_FMT, 14: METRIC_FMT, 15: METRIC_FMT, 16: METRIC_FMT, 17: METRIC_FMT, 18: METRIC_FMT,
        19: METRIC_FMT, 20: METRIC_FMT, 21: METRIC_FMT,
    }
    last_row = write_table(ws, start_row, headers, data_rows, col_formats)

    n_cols = len(headers)
    f1_col = 12  # "F1"
    india_auc_col = 19
    italy_auc_col = 20
    gap_col = 21
    f1_letter = get_column_letter(f1_col)
    india_letter = get_column_letter(india_auc_col)
    italy_letter = get_column_letter(italy_auc_col)
    f1_range = f"${f1_letter}${start_row + 1}:${f1_letter}${last_row}"

    for i, r in enumerate(rows):
        excel_row = start_row + 1 + i
        # RANK() (not RANK.EQ) -- the classic pre-2007 function, since this
        # environment has no LibreOffice available to verify a newer
        # function's evaluation (see build script's own run notes / the
        # chat response this was built from). RANK() is fully supported by
        # every Excel version and gives an identical result to RANK.EQ for
        # non-tied values (ties, if any, would both get the same rank under
        # either function -- not a behavioral difference here).
        ws.cell(row=excel_row, column=1, value=f"=RANK({f1_letter}{excel_row},{f1_range},0)")
        ws.cell(row=excel_row, column=1).font = BODY_FONT
        ws.cell(row=excel_row, column=1).alignment = CENTER
        ws.cell(row=excel_row, column=1).border = THIN_BORDER

        gap_cell = ws.cell(row=excel_row, column=gap_col)
        if r["auc_india"] is not None and r["auc_italy"] is not None:
            gap_cell.value = f"=ABS({india_letter}{excel_row}-{italy_letter}{excel_row})"
        gap_cell.font = BODY_FONT
        gap_cell.number_format = METRIC_FMT
        gap_cell.alignment = CENTER
        gap_cell.border = THIN_BORDER

        # Flag the known collapsed combo (08_new_way_architecture_roster.md
        # SS7: best trial = trial 0, near-chance accuracy) so it isn't
        # quietly read as a normal result in a sorted table.
        if r["model_name"] == "convnext_large_forniceal_palpebral_new_way":
            for c in range(1, n_cols + 1):
                ws.cell(row=excel_row, column=c).fill = FLAG_FILL

    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)
    ws.auto_filter.ref = f"A{start_row}:{get_column_letter(n_cols)}{last_row}"

    note_row = last_row + 2
    ws.cell(row=note_row, column=1, value="Row highlighted red: best trial = trial 0 (no improvement across the other 11 trials), accuracy 0.548 -- looks collapsed, not just weak. See 08_new_way_architecture_roster.md §8 before citing.").font = SUBTITLE_FONT

    widths = {"A": 6, "B": 40, "C": 20, "D": 9, "E": 9, "F": 11, "G": 9, "H": 11,
              "I": 13, "J": 13, "K": 12, "L": 8, "M": 9, "N": 9, "O": 9, "P": 10, "Q": 15, "R": 8, "S": 11, "T": 11, "U": 15}
    autofit(ws, widths)

    # ---- Sheet 2: Per-Country Detail ----
    ws2 = wb.create_sheet("Per-Country Detail")
    ws2["A1"] = "Per-Country Stratified Metrics (India vs. Italy)"
    ws2["A1"].font = TITLE_FONT
    ws2["A2"] = "Same 16 combos as Summary. n = validation patients in that country's slice."
    ws2["A2"].font = SUBTITLE_FONT

    headers2 = [
        "Model", "Tissue Type",
        "N (India)", "Accuracy (India)", "Precision (India)", "Recall (India)", "Specificity (India)", "Balanced Acc (India)", "F1 (India)", "AUC (India)", "TN/FP/FN/TP (India)",
        "N (Italy)", "Accuracy (Italy)", "Precision (Italy)", "Recall (Italy)", "Specificity (Italy)", "Balanced Acc (Italy)", "F1 (Italy)", "AUC (Italy)", "TN/FP/FN/TP (Italy)",
    ]
    data_rows2 = []
    for r in rows:
        cm_i = f"{r['tn_india']}/{r['fp_india']}/{r['fn_india']}/{r['tp_india']}" if r["tn_india"] is not None else "N/A"
        cm_it = f"{r['tn_italy']}/{r['fp_italy']}/{r['fn_italy']}/{r['tp_italy']}" if r["tn_italy"] is not None else "N/A"
        data_rows2.append(
            [
                r["model_name"], r["tissue_type"],
                r["n_india"], r["accuracy_india"], r["precision_india"], r["recall_india"], r["specificity_india"], r["balanced_accuracy_india"], r["f1_india"], r["auc_india"], cm_i,
                r["n_italy"], r["accuracy_italy"], r["precision_italy"], r["recall_italy"], r["specificity_italy"], r["balanced_accuracy_italy"], r["f1_italy"], r["auc_italy"], cm_it,
            ]
        )
    col_formats2 = {c: METRIC_FMT for c in [4, 5, 6, 7, 8, 9, 10, 13, 14, 15, 16, 17, 18, 19]}
    last_row2 = write_table(ws2, 4, headers2, data_rows2, col_formats2)
    ws2.freeze_panes = ws2.cell(row=5, column=1)
    ws2.auto_filter.ref = f"A4:{get_column_letter(len(headers2))}{last_row2}"
    widths2 = {"A": 40, "B": 20}
    for col in "CDEFGHIJK":
        widths2[col] = 12
    for col in "LMNOPQRST":
        widths2[col] = 12
    autofit(ws2, widths2)

    # ---- Sheet 3: Hyperparameters ----
    ws3 = wb.create_sheet("Hyperparameters")
    ws3["A1"] = "Best Hyperparameters per Model"
    ws3["A1"].font = TITLE_FONT
    ws3["A2"] = "Optuna-tuned per trial: learning_rate ~ LogUniform(1e-4, 1e-1), weight_decay ~ LogUniform(1e-6, 1e-3), dropout_rate in {0.2, 0.5}. Values shown are the winning trial's."
    ws3["A2"].font = SUBTITLE_FONT

    headers3 = ["Model", "Tissue Type", "Architecture Family", "N Trials Run", "Best Trial #", "Learning Rate", "Weight Decay", "Dropout Rate"]
    data_rows3 = [
        [r["model_name"], r["tissue_type"], r["family"], r["n_trials_run"], r["best_trial_number"], r["learning_rate"], r["weight_decay"], r["dropout_rate"]]
        for r in sorted(rows, key=lambda r: r["model_name"])
    ]
    col_formats3 = {6: SCI_FMT, 7: SCI_FMT, 8: "0.00"}
    last_row3 = write_table(ws3, 4, headers3, data_rows3, col_formats3)
    ws3.freeze_panes = ws3.cell(row=5, column=1)
    ws3.auto_filter.ref = f"A4:{get_column_letter(len(headers3))}{last_row3}"
    autofit(ws3, {"A": 42, "B": 20, "C": 18, "D": 11, "E": 11, "F": 13, "G": 13, "H": 12})

    # ---- Sheet 4: Notes ----
    ws4 = wb.create_sheet("Notes")
    ws4["A1"] = "Notes & Provenance"
    ws4["A1"].font = TITLE_FONT
    notes = [
        "",
        "Data source: classification/new_way/Output/logs/*_study_summary.json, read directly (16 files, one per combo).",
        "This is the real Kaggle run committed 2026-08-11 (commit 4f89390 \"v2_clean_scripts is added\") -- see classification/.project_memory/08_new_way_architecture_roster.md §8.",
        "",
        "All metrics are from the VALIDATION split, at the single best trial (by validation F1) of each combo's 12-trial Optuna search -- not the held-out test set.",
        "",
        "\"sensitivity\" and \"recall\" are the same quantity for the anemic (positive) class -- only \"Recall\" is shown to avoid a duplicate column.",
        "",
        "Known issue: convnext_large_forniceal_palpebral_new_way (flagged red on the Summary sheet) has its best trial at trial 0 -- no improvement found across the remaining 11 trials -- and accuracy 0.548 (near chance), with 14 false positives against only 3 true negatives. This looks like a collapsed run, not a genuinely weak one. Not yet root-caused. Do not cite this combo's numbers without investigating further.",
        "",
        "Architecture family/tier/parameter counts are cited from classification/.project_memory/08_new_way_architecture_roster.md §2 (verified by constructing each model and summing p.numel()), not re-measured for this workbook.",
        "",
        "India/Italy AUC Gap = |AUC(India) - AUC(Italy)|, computed by formula in this workbook, not hardcoded. This project treats a large gap as evidence a model may be exploiting the India/Italy demographic confound (root CLAUDE.md §0.5) rather than genuine conjunctival pallor -- see classification/.project_memory/07_step1_measurement_harness.md for the fuller defensibility analysis, and note that project's own caveat that per-model AUC rankings at this sample size (India val n=14, Italy val n=17-19) are noisy and should not be over-interpreted individually.",
        "",
        "Generated by build_comparison_excel.py in this same folder -- re-run it to refresh this workbook if the underlying study_summary.json files change.",
    ]
    for i, line in enumerate(notes, start=2):
        cell = ws4.cell(row=i, column=1, value=line)
        cell.font = BODY_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws4.column_dimensions["A"].width = 130

    return wb


def main():
    rows = load_combos()
    print(f"Loaded {len(rows)} combos from {LOGS_DIR}")
    wb = build_workbook(rows)
    OUTPUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_XLSX)
    print(f"Saved {OUTPUT_XLSX}")


if __name__ == "__main__":
    main()
